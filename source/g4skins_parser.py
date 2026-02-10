from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
import os
import pandas as pd
import re

# =========================
# CONFIG
# =========================

DEBUGGER_ADDRESS = "127.0.0.1:9222"
EXCEL_PATH = "g4skins.xlsx"

QUALITIES = {"BS", "WW", "FT", "MW", "FN"}

wear_map = {
    "FN": "Factory New",
    "MW": "Minimal Wear",
    "FT": "Field-Tested",
    "WW": "Well-Worn",
    "BS": "Battle-Scarred"
}

def clean_name(name: str) -> str:
    if not isinstance(name, str):
        return name

    name = re.sub(r"\s+", " ", name.strip())

    wear_found = None
    for short, full in wear_map.items():
        pattern = rf"(?:\(\s*{short}\s*\)|\b{short}\b)"
        if re.search(pattern, name):
            wear_found = full
            name = re.sub(pattern, "", name).strip()
            break

    if re.search(r"\bST\b", name):
        name = re.sub(r"\bST\b", "", name).strip()
        name = f"StatTrak™ {name}"

    if wear_found:
        name = f"{name} ({wear_found})"

    return name

# =========================
# CONNECT TO OPEN CHROME
# =========================

options = Options()
options.add_experimental_option("debuggerAddress", DEBUGGER_ADDRESS)

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wait = WebDriverWait(driver, 15)
actions = ActionChains(driver)

# =========================
# EXCEL
# =========================

if os.path.exists(EXCEL_PATH):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    print(f"📘 Opened existing file: {EXCEL_PATH}")
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["steam_market_hash_name", "g4skins_price"])
    print(f"🆕 Created new file: {EXCEL_PATH}")

existing = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    existing.add(r[0])

# =========================
# MAIN PAGE — GET CASES
# =========================

wait.until(
    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a.g_case.CaseBox"))
)

case_links = driver.find_elements(By.CSS_SELECTOR, "a.g_case.CaseBox")
case_urls = [c.get_attribute("href") for c in case_links]

print(f"🔍 Found {len(case_urls)} cases")

# =========================
# PROCESS EACH CASE
# =========================

for idx, case_url in enumerate(case_urls, start=1):
    print(f"\n➡️ [{idx}] Opening case: {case_url}")
    driver.get(case_url)

    # ----- wait items -----
    wait.until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".list-item"))
    )

    # ----- case name -----
    case_name = driver.find_element(By.CSS_SELECTOR, "h1").text.strip()
    print(f"📦 Case: {case_name}")

    items = driver.find_elements(By.CSS_SELECTOR, ".list-item")

    if not items:
        print("⚠️ No items found, skipping case")
        continue

    # ----- ONE HOVER (hydration) -----
    first_item = items[0]
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", first_item
    )
    time.sleep(0.6)

    hover_target = first_item.find_element(By.CSS_SELECTOR, ".content-image")
    actions.move_to_element(hover_target).pause(1.2).perform()
    time.sleep(2.0)

    # ----- PARSE VIA BEAUTIFULSOUP -----
    soup = BeautifulSoup(driver.page_source, "html.parser")
    soup_items = soup.select(".list-item")

    for item in soup_items:
        weapon_el = item.select_one(
            ".G_Text.bottom-weapon .G_Text-content"
        )
        skin_el = item.select_one(
            ".G_Text.bottom-skin .G_Text-content"
        )

        weapon = weapon_el.text.strip() if weapon_el else ""
        skin = skin_el.text.strip() if skin_el else ""

        if not weapon or not skin:
            continue

        rows = item.select(".content-hover tbody tr")

        for row in rows:
            texts = [
                p.text.strip()
                for p in row.select("p.G_Text-content")
                if p.text.strip()
            ]

            quality = next((t for t in texts if t in QUALITIES), "")
            price = next((t.replace("$", "") for t in texts if "$" in t), "")

            if not quality or not price:
                continue

            raw_name  = f"{weapon} | {skin} ({quality})"
            full_name = clean_name(raw_name)
            key = full_name

            if key in existing:
                continue

            ws.append([full_name, price])
            existing.add(key)

            print(f"   ✔ {full_name} = {price}")

    wb.save(EXCEL_PATH)
    print(f"💾 Saved case: {case_name}")

# =========================
# DONE
# =========================

print("\n🎉 ALL CASES DONE")
driver.quit()

def save_distinct_g4skins():
    print("🧹 Deduplicating by steam_market_hash_name...")
    df = pd.read_excel(EXCEL_PATH)

    before = len(df)

    df = (
        df
        .dropna(subset=["steam_market_hash_name"])
        .drop_duplicates(
            subset=["steam_market_hash_name"],
            keep="first"
        )
    )

    after = len(df)

    df.to_excel(EXCEL_PATH, index=False)

    print(f"✅ Dedup done: {before} → {after}")

def merge_with_problematic():
    PROBLEMATIC_FILE = "Problematic Withdrawals.xlsx"

    print("🔗 Merging g4skins_price into Problematic Withdrawals.xlsx ...")

    # 1. Lookup з g4skins
    g4_df = pd.read_excel(EXCEL_PATH)
    g4_df = (
        g4_df
        .dropna(subset=["steam_market_hash_name"])
        .drop_duplicates("steam_market_hash_name", keep="first")
    )

    price_map = (
        g4_df
        .set_index("steam_market_hash_name")["g4skins_price"]
        .to_dict()
    )

    # 2. Всі аркуші Problematic
    sheets = pd.read_excel(PROBLEMATIC_FILE, sheet_name=None)
    out_sheets = {}

    for sheet_name, sheet_df in sheets.items():
        print(f"   🔄 Processing sheet: {sheet_name}")

        if "steam_market_hash_name" not in sheet_df.columns:
            print("   ⚠️ Skipped (no steam_market_hash_name)")
            out_sheets[sheet_name] = sheet_df
            continue

        # 3. overwrite або create
        sheet_df["g4skins_price"] = (
            sheet_df["steam_market_hash_name"]
            .map(price_map)
            .fillna("-")
        )

        out_sheets[sheet_name] = sheet_df

    # 4. Запис назад усіх аркушів
    with pd.ExcelWriter(
        PROBLEMATIC_FILE,
        engine="openpyxl",
        mode="w"
    ) as writer:
        for sheet_name, df_out in out_sheets.items():
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)

    print("✅ g4skins_price overwritten / created in all applicable sheets")

def main():
    print("\n=== SELECT ACTION ===")
    print("1. Save distinct g4skins.xlsx")
    print("2. Merge with Problematic Withdrawals.xlsx")

    choice = input("\nEnter choice (1/2): ").strip()

    if choice == "1":
        save_distinct_g4skins()
    elif choice == "2":
        # save_distinct_g4skins()  # ⬅️ гарантія чистих даних
        merge_with_problematic()
    else:
        print("❌ Invalid choice")

if __name__ == "__main__":
    main()

# run:
# cmd
# "C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\ChromeDebug"
# and visit main site before running the script