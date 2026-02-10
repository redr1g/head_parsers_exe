import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bs4 import BeautifulSoup
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
import time
import os
import re

# =========================
# CONFIG
# =========================

DEBUGGER_ADDRESS = "127.0.0.1:9222"
BASE_URL = "https://csgo-skins.com"
EXCEL_PATH = "csgoskins.xlsx"

wear_map = {
    "FN": "Factory New",
    "MW": "Minimal Wear",
    "FT": "Field-Tested",
    "WW": "Well-Worn",
    "BS": "Battle-Scarred"
}

star_keywords = [
    "Knife", "Gloves", "Karambit",
    "Bayonet", "Shadow Daggers", "Hand Wraps"
]

def clean_name(name: str) -> str:
    if not isinstance(name, str):
        return name

    # 1. Прибрати подвійні пробіли
    name = re.sub(r"\s+", " ", name.strip())

    # 2. Перевірити наявність wear-стану (FN, MW, FT, WW, BS)
    wear_found = None
    for short, full in wear_map.items():
        if re.search(rf"\b{short}\b", name):
            wear_found = full
            name = re.sub(rf"\b{short}\b", "", name).strip()
            break

    # 3. Замінити ST на StatTrak™ (на початку назви)
    if re.search(r"\bST\b", name):
        name = re.sub(r"\bST\b", "", name).strip()
        name = f"StatTrak™ {name}"

    # 4. Якщо в назві є knife/gloves — додати ★
    if any(keyword.lower() in name.lower() for keyword in star_keywords):
        if not name.lstrip().startswith("★"):
            name = f"★ {name}"

    # 5. Додати wear-стан у дужках, якщо був знайдений
    if wear_found:
        name = f"{name} ({wear_found})"

    return name

# =========================
# CONNECT TO OPEN CHROME
# =========================

options = Options()
options.add_experimental_option("debuggerAddress", DEBUGGER_ADDRESS)

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wait = WebDriverWait(driver, 20)
actions = ActionChains(driver)

# =========================
# EXCEL
# =========================

if os.path.exists(EXCEL_PATH):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    print(f"📘 Відкрито існуючий файл: {EXCEL_PATH}")
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["steam_market_hash_name", "csgoskins_price"])
    print(f"🆕 Створено новий файл: {EXCEL_PATH}")

existing_entries = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    existing_entries.add(row[0])

# =========================
# OPEN MAIN PAGE
# =========================

driver.get(BASE_URL)

wait.until(
    EC.presence_of_all_elements_located(
        (By.CSS_SELECTOR, "article.ContainersContainer")
    )
)

print("🏠 Головна сторінка завантажена")

# =========================
# GET CASE COUNT
# =========================

cases = driver.find_elements(
    By.CSS_SELECTOR,
    "article.ContainersContainer"
)
total_cases = len(cases)

print(f"🔍 Знайдено кейсів: {total_cases}")

# =========================
# ITERATE CASES
# =========================
def parse_rare_modal():
    rows = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, ".ContainerChancesModal_table tbody tr")
        )
    )

    for r in rows:
        name = r.find_element(
            By.CSS_SELECTOR,
            ".ContainerChancesModal_item-name"
        ).text.strip()

        cells = r.find_elements(By.CSS_SELECTOR, "td.table_cell")
        if len(cells) < 2:
            continue

        price = cells[1].text.strip().replace("$", "")

        if name in existing_entries:
            continue

        ws.append([name, price])
        existing_entries.add(name)
        print(f"{name} = {price}")

    # close modal
    driver.find_element(By.CSS_SELECTOR, ".Modal_close").click()
    time.sleep(0.6)

for case_index in range(total_cases):
    print(f"\n➡️ Відкриваю кейс [{case_index + 1}/{total_cases}]")

    # ЗАВЖДИ перевитягуємо кейси (DOM міняється)
    cases = driver.find_elements(
        By.CSS_SELECTOR,
        "article.ContainersContainer"
    )
    case = cases[case_index]

    # Людський клік
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", case
    )
    time.sleep(0.8)

    actions.move_to_element(case).pause(0.4).click().perform()

    # =========================
    # WAIT CASE PAGE
    # =========================

    wait.until(
        EC.presence_of_all_elements_located(
            (By.CLASS_NAME, "list_item")
        )
    )
    time.sleep(1.5)

    print("📦 Сторінка кейса відкрита")

    # =========================
    # ====== ТВОЙ СТАРИЙ КОД ======
    # =========================

    list_items = driver.find_elements(By.CLASS_NAME, "list_item")
    print(f"🔍 Found {len(list_items)} items on page.")

    for item in list_items:
        try:
            html = item.get_attribute("innerHTML")
            soup = BeautifulSoup(html, "html.parser")

            base_name_tag = soup.select_one(".ContainerGroupedItem_name")
            if not base_name_tag:
                continue
            base_name = base_name_tag.text.strip()

            # =========================
            # RARE SPECIAL ITEM HANDLER
            # =========================
            if "Rare Special Item" in base_name:
                try:
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", item
                    )
                    time.sleep(0.4)

                    actions.move_to_element(item).pause(0.3).click().perform()

                    parse_rare_modal()
                    continue

                except Exception as e:
                    print(f"❌ Rare modal error: {e}")
                    continue
                
            rows = soup.select("table.chances_table tbody tr")
            for row in rows:
                stattrak_td = row.select_one(".cell--is-statTrak")
                condition_td = row.select_one(
                    "td:not(.cell--is-statTrak):not(.cell--text-primary-color)"
                )

                if stattrak_td:
                    variant_name = stattrak_td.text.strip()
                    full_name = clean_name(f"{base_name} {variant_name}")
                elif condition_td:
                    variant_name = condition_td.text.strip()
                    full_name = clean_name(f"{base_name} {variant_name}")
                else:
                    full_name = clean_name(base_name)

                price_td = row.select_one(
                    ".cell--text-primary-color .Currency"
                )
                if not price_td:
                    continue

                price_text = price_td.text.strip().replace("$", "")

                if full_name in existing_entries:
                    continue

                ws.append([full_name, price_text])
                existing_entries.add(full_name)

                print(f"   ✔ {full_name} = {price_text}")

        except Exception as e:
            print(f"❌ Error: {e}")

    wb.save(EXCEL_PATH)
    print("💾 Кейc збережено")

    # =========================
    # BACK TO MAIN PAGE
    # =========================

    driver.back()

    wait.until(
        EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "article.ContainersContainer")
        )
    )
    time.sleep(1.2)

# =========================
# DONE
# =========================

print("\n🎉 ВСІ КЕЙСИ ОБРОБЛЕНО")
driver.quit()

def save_distinct_csgoskins():
    print("🧹 Deduplicating by steam_market_hash_name...")
    df = pd.read_excel(EXCEL_PATH)

    before = len(df)

    df = (
        df
        .dropna(subset=["steam_market_hash_name"])
        .drop_duplicates(
            subset=["steam_market_hash_name"],
            keep="first"
        )
    )

    after = len(df)

    df.to_excel(EXCEL_PATH, index=False)

    print(f"✅ Dedup done: {before} → {after}")


def merge_with_problematic():
    PROBLEMATIC_FILE = "Problematic Withdrawals.xlsx"

    print("🔗 Merging csgoskins_price into Problematic Withdrawals.xlsx ...")

    # 1. Lookup з g4skins
    g4_df = pd.read_excel(EXCEL_PATH)
    g4_df = (
        g4_df
        .dropna(subset=["steam_market_hash_name"])
        .drop_duplicates("steam_market_hash_name", keep="first")
    )

    price_map = (
        g4_df
        .set_index("steam_market_hash_name")["csgoskins_price"]
        .to_dict()
    )

    # 2. Всі аркуші Problematic
    sheets = pd.read_excel(PROBLEMATIC_FILE, sheet_name=None)
    out_sheets = {}

    for sheet_name, sheet_df in sheets.items():
        print(f"   🔄 Processing sheet: {sheet_name}")

        if "steam_market_hash_name" not in sheet_df.columns:
            print("   ⚠️ Skipped (no steam_market_hash_name)")
            out_sheets[sheet_name] = sheet_df
            continue

        # 3. overwrite або create
        sheet_df["csgoskins_price"] = (
            sheet_df["steam_market_hash_name"]
            .map(price_map)
            .fillna("-")
        )

        out_sheets[sheet_name] = sheet_df

    # 4. Запис назад усіх аркушів
    with pd.ExcelWriter(
        PROBLEMATIC_FILE,
        engine="openpyxl",
        mode="w"
    ) as writer:
        for sheet_name, df_out in out_sheets.items():
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)

    print("✅ csgoskins_price overwritten / created in all applicable sheets")


def main():
    print("\n=== SELECT ACTION ===")
    print("1. Save distinct g4skins.xlsx")
    print("2. Merge with Problematic Withdrawals.xlsx")

    choice = input("\nEnter choice (1/2): ").strip()

    if choice == "1":
        save_distinct_csgoskins()
    elif choice == "2":
        # save_distinct_g4skins()  # ⬅️ гарантія чистих даних
        merge_with_problematic()
    else:
        print("❌ Invalid choice")

if __name__ == "__main__":
    main()